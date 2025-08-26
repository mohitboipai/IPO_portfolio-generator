import pandas as pd
import itertools
from openpyxl import Workbook

def read_ipos(file_path):
    # Read the CSV file into a DataFrame
    df = pd.read_csv(file_path)
    return df

def filter_unique_ipos(df):
    # Filter unique IPOs based on a unique identifier (e.g., 'ticker')
    return df.drop_duplicates(subset='ticker')

def generate_portfolios(df):
    portfolios = []
    # Generate all combinations of IPOs
    for r in range(1, len(df) + 1):
        for combination in itertools.combinations(df.iterrows(), r):
            # Check for overlapping listing dates
            if not has_overlapping_dates(combination):
                portfolios.append(combination)
    return portfolios

def has_overlapping_dates(combination):
    # Check if any IPOs in the combination have overlapping listing dates
    dates = []
    for index, row in combination:
        dates.append((row['listing_date'], row['listing_date']))  # Assuming listing_date is in the format 'YYYY-MM-DD'
    
    # Sort dates and check for overlaps
    dates.sort()
    for i in range(1, len(dates)):
        if dates[i][0] <= dates[i-1][1]:  # Overlap condition
            return True
    return False

def analyze_portfolio(portfolio):
    # Perform analysis on the portfolio
    analysis = {
        'total_ipos': len(portfolio),
        'total_market_cap': sum(row['market_cap'] for index, row in portfolio)  # Assuming market_cap is a column
    }
    return analysis

def save_to_excel(portfolios, output_file):
    wb = Workbook()
    for i, portfolio in enumerate(portfolios):
        ws = wb.create_sheet(title=f'Portfolio {i + 1}')
        for index, row in portfolio:
            ws.append(row.tolist())  # Append the row data
        analysis = analyze_portfolio(portfolio)
        ws.append(['Analysis'])
        ws.append(['Total IPOs', analysis['total_ipos']])
        ws.append(['Total Market Cap', analysis['total_market_cap']])
    
    # Remove the default sheet created by Workbook
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    wb.save(output_file)

def main():
    file_path = 'ipos.csv'  # Path to your CSV file
    output_file = 'portfolios.xlsx'  # Output Excel file
    df = read_ipos(file_path)
    unique_ipos = filter_unique_ipos(df)
    portfolios = generate_portfolios(unique_ipos)
    save_to_excel(portfolios, output_file)
    print(f'Saved {len(portfolios)} portfolios to {output_file}')

if __name__ == "__main__":
    main()