import pandas as pd
from itertools import combinations
from openpyxl import Workbook

def read_ipos(file_path):
    # Read the CSV file into a DataFrame
    return pd.read_csv(file_path)

def filter_unique_ipos(df):
    # Assuming 'IPO_ID' is the unique identifier for IPOs
    return df.drop_duplicates(subset=['IPO_ID'])

def generate_portfolios(df):
    portfolios = []
    n = len(df)
    
    # Generate all combinations of IPOs
    for r in range(1, n + 1):
        for combo in combinations(df.iterrows(), r):
            # Check for overlapping listing dates
            if not has_overlapping_dates(combo):
                portfolios.append(combo)
    
    return portfolios

def has_overlapping_dates(combo):
    # Extract listing dates from the combination
    listing_dates = [row[1]['Listing_Date'] for row in combo]
    return len(listing_dates) != len(set(listing_dates))

def analyze_portfolio(combo):
    # Perform analysis on the portfolio
    total_market_cap = sum(row[1]['Market_Cap'] for row in combo)
    return {
        'Number of IPOs': len(combo),
        'Total Market Cap': total_market_cap
    }

def save_to_excel(portfolios, output_file):
    wb = Workbook()
    
    for i, combo in enumerate(portfolios):
        sheet = wb.create_sheet(title=f'Portfolio {i + 1}')
        
        # Write IPO data to the sheet
        for j, row in enumerate(combo):
            sheet.append(row[1].tolist())
        
        # Add analysis
        analysis = analyze_portfolio(combo)
        sheet.append(['Analysis'])
        for key, value in analysis.items():
            sheet.append([key, value])
    
    # Remove the default sheet created by Workbook
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    wb.save(output_file)

def main():
    file_path = 'ipos.csv'  # Path to your CSV file
    output_file = 'portfolios.xlsx'  # Output Excel file
    
    # Step 1: Read IPOs
    ipos_df = read_ipos(file_path)
    
    # Step 2: Filter unique IPOs
    unique_ipos_df = filter_unique_ipos(ipos_df)
    
    # Step 3: Generate portfolios
    portfolios = generate_portfolios(unique_ipos_df)
    
    # Step 4: Save to Excel
    save_to_excel(portfolios, output_file)

if __name__ == "__main__":
    main()